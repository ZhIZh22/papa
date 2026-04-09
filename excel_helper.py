import os
import re
import shutil
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

EXCEL_FILE = os.getenv("EXCEL_FILE", "!!!Мастерская.xlsx")

# Тонкая боковая граница (как на скриншоте)
_THIN = Side(style="thin")
_BORDER_SIDES = Border(left=_THIN, right=_THIN)  # только левая и правая полоски
_BORDER_COMMENT = Border(left=_THIN, right=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")


def _apply_row_style(ws, row_idx: int):
    """Применяет стиль (выравнивание + боковые границы) к строке данных."""
    for col in range(1, 5):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = _BORDER_SIDES
        if col == 4:
            cell.alignment = _ALIGN_LEFT
        else:
            cell.alignment = _ALIGN_CENTER


def _get_sheet_name(year: int) -> str:
    return str(year)[-2:]


def _format_money(value) -> str:
    if value is None:
        return ""
    try:
        n = int(value)
    except (ValueError, TypeError):
        return str(value)
    result = []
    s = str(abs(n))
    for i, ch in enumerate(reversed(s)):
        if i > 0 and i % 3 == 0:
            result.append(" ")
        result.append(ch)
    formatted = "".join(reversed(result))
    return ("-" + formatted) if n < 0 else formatted


def _parse_money(s):
    if s is None:
        return None
    s = str(s).replace(" ", "").replace("\xa0", "")
    try:
        return int(s)
    except ValueError:
        return None


def _ensure_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = "Дата"
        ws["B1"] = "Приход"
        ws["C1"] = "Расход"
        ws["D1"] = "Комментарий"
        ws["A2"] = "Итого:"
        ws["A3"] = "Остаток:"
    return wb[sheet_name]



def _last_data_row(ws) -> int:
    """Возвращает номер последней строки с датой (>= 4), или 3 если данных нет."""
    last = 3
    for row_idx in range(4, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value:
            last = row_idx
    return last


def _last_date_in_sheet(ws) -> date | None:
    """Возвращает последнюю дату из столбца A (строки >= 4)."""
    last_date = None
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val:
            d = _parse_date(str(val).strip())
            if d and (last_date is None or d > last_date):
                last_date = d
    return last_date


def _parse_date(s: str) -> date | None:
    for fmt in ("%d.%m.%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _find_row_by_date(ws, target_date: date) -> int | None:
    """Ищет строку с нужной датой. Возвращает номер или None."""
    date_str = target_date.strftime("%d.%m.%y")
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip() == date_str:
            return row_idx
    return None


def _find_or_create_date_row(ws, target_date: date) -> int:
    """
    Находит строку с нужной датой или вставляет новую в правильное место.
    - Дата позже последней: добавляет пустые строки для каждого пропущенного дня.
    - Дата раньше последней: вставляет строку между соседними датами.
    - Нет данных: просто добавляет строку после строки 3.
    """
    # Уже есть?
    existing = _find_row_by_date(ws, target_date)
    if existing:
        return existing

    # Собираем все строки с датами: {номер_строки: date}
    rows_with_dates = {}
    for row_idx in range(4, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val:
            d = _parse_date(str(val).strip())
            if d:
                rows_with_dates[row_idx] = d

    if not rows_with_dates:
        # Нет данных — первая запись
        ws.cell(row=4, column=1).value = target_date.strftime("%d.%m.%y")
        _apply_row_style(ws, 4)
        return 4

    last_row  = _last_data_row(ws)
    last_date = max(rows_with_dates.values())

    # Для любой даты (прошлое или будущее) — ищем ближайшую строку ДО target_date
    # и вставляем новую строку сразу после неё.
    insert_after_row = 3  # по умолчанию сразу после заголовков
    for row_idx in sorted(rows_with_dates.keys()):
        if rows_with_dates[row_idx] < target_date:
            insert_after_row = row_idx

    if target_date > last_date:
        # Дата позже всех — добавляем в конец (не вставляем, чтобы не сдвигать)
        new_row = last_row + 1
        ws.cell(row=new_row, column=1).value = target_date.strftime("%d.%m.%y")
        _apply_row_style(ws, new_row)
        return new_row

    # Дата в середине или в начале — вставляем строку
    insert_at = insert_after_row + 1
    ws.insert_rows(insert_at)
    ws.cell(row=insert_at, column=1).value = target_date.strftime("%d.%m.%y")
    _apply_row_style(ws, insert_at)
    return insert_at


def add_transaction(target_date: date, amount: int, comment: str = "") -> dict:
    wb = load_workbook(EXCEL_FILE)
    sheet_name = _get_sheet_name(target_date.year)
    ws = _ensure_sheet(wb, sheet_name)

    row_idx = _find_or_create_date_row(ws, target_date)

    if amount > 0:
        existing = _parse_money(ws.cell(row=row_idx, column=2).value) or 0
        ws.cell(row=row_idx, column=2).value = existing + amount
    else:
        existing = _parse_money(ws.cell(row=row_idx, column=3).value) or 0
        ws.cell(row=row_idx, column=3).value = existing + abs(amount)

    if comment:
        existing_comment = ws.cell(row=row_idx, column=4).value or ""
        if existing_comment:
            ws.cell(row=row_idx, column=4).value = existing_comment + ", " + comment
        else:
            ws.cell(row=row_idx, column=4).value = comment

    # Применяем стиль к строке
    _apply_row_style(ws, row_idx)

    wb.save(EXCEL_FILE)
    return get_totals(target_date.year)


def get_totals(year: int) -> dict:
    wb = load_workbook(EXCEL_FILE)
    sheet_name = _get_sheet_name(year)
    if sheet_name not in wb.sheetnames:
        return {"income": 0, "expense": 0, "balance": 0}
    ws = wb[sheet_name]
    total_income = 0
    total_expense = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        income  = _parse_money(row[1]) if row[1] else 0
        expense = _parse_money(row[2]) if row[2] else 0
        if income:
            total_income  += income
        if expense:
            total_expense += expense
    return {"income": total_income, "expense": total_expense, "balance": total_income - total_expense}


def get_day_info(target_date: date) -> dict:
    wb = load_workbook(EXCEL_FILE)
    sheet_name = _get_sheet_name(target_date.year)
    if sheet_name not in wb.sheetnames:
        return {"date": target_date, "income": 0, "expense": 0, "comment": ""}
    ws = wb[sheet_name]
    row_idx = _find_row_by_date(ws, target_date)
    if row_idx:
        income  = _parse_money(ws.cell(row=row_idx, column=2).value) or 0
        expense = _parse_money(ws.cell(row=row_idx, column=3).value) or 0
        comment = ws.cell(row=row_idx, column=4).value or ""
        return {"date": target_date, "income": income, "expense": expense, "comment": comment}
    return {"date": target_date, "income": 0, "expense": 0, "comment": ""}


def replace_excel_file(new_file_path: str):
    shutil.copy2(new_file_path, EXCEL_FILE)
